import openpyxl
from openpyxl import load_workbook
import os


def loading_file(filepath):
    try:
        workbook = load_workbook(filepath, rich_text=True)
        print(f'[INFO] Loading successfull')
        print(f'[INFO] Path of Workbook: {filepath}')
        print(f'[INFO] Name of Workbook: {os.path.basename(filepath)}')
        return workbook
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading failed.')
        pass
    

def load_DataSheet(workbook):
    try:
        for name in workbook.sheetnames:
            if name == 'Data Sheet':
                print(f'[INFO] "Data Sheet" found in workbook')
                sheet_DataSheet = workbook.active
                return sheet_DataSheet
            else:
                print(f'[INFO] "Data Sheet" not found in workbook')
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading failed.')
        pass
       
def get_company_name(Worksheet):
    try:
        company_name = Worksheet.cell(column=2,row=1).value
        print(f'[INFO] Company name found')
        return company_name
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading name failed.')
        pass
'''
def save_cells_to_dict(worksheet, start_cell, end_cell):
    cell_dict = {}
    
    # Extract values from each cell in the specified range
    for row in worksheet[start_cell:end_cell]:
        for cell in row:
            cell_dict[cell.coordinate] = cell.value
    
    # Print the dictionary (for demonstration purposes)
    print(cell_dict)
'''

def save_metadata_to_dict(worksheet, start_cell, end_cell):
    try:
        # defining a dictionary
        metadata_dict = {}
        #iterating over each row
        for row in worksheet[start_cell:end_cell]:
            key_cell, value_cell = row
            metadata_dict[key_cell.value] = value_cell.value
        
        print(f'[INFO] saved metadata into dictionary')
        # Print the dictionary (for demonstration purposes)
        print(metadata_dict)
        # returning the dictionary value
        return metadata_dict
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading metadata failed.')
        pass
    
def save_profit_loss_to_dict(worksheet, start_cell, end_cell):
    try:
        profit_loss_data_dict = {}
        
        # Extract values from each cell in the specified range
        for row in worksheet[start_cell:end_cell]:
            key = row[0].value  # Extract the key from the first cell of the row
            values = [cell.value for cell in row[1:]]  # Extract values from the remaining cells of the row
            profit_loss_data_dict[key] = values
            
        print('[INFO] saved profit and loss in dictionary')
        print(profit_loss_data_dict)
        return profit_loss_data_dict
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading profit and loss failed.')
        pass
    
def save_quarters_to_dict(worksheet, start_cell, end_cell):
    try:
        quarters_data_dict = {}
        
        # Extract values from each cell in the specified range
        for row in worksheet[start_cell:end_cell]:
            key = row[0].value  # Extract the key from the first cell of the row
            values = [cell.value for cell in row[1:]]  # Extract values from the remaining cells of the row
            quarters_data_dict[key] = values
            
        print('[INFO] saved quarters to dictionary')
        print(quarters_data_dict)
        return quarters_data_dict
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading quarters failed.')
        pass
    
def save_balance_sheet_to_dict(worksheet, start_cell, end_cell):
    try:
        balance_sheet_data_dict = {}
        
        # Extract values from each cell in the specified range
        for row in worksheet[start_cell:end_cell]:
            key = row[0].value  # Extract the key from the first cell of the row
            values = [cell.value for cell in row[1:]]  # Extract values from the remaining cells of the row
            balance_sheet_data_dict[key] = values
            
        print('[INFO] saved balance sheet to dictionary')
        print(balance_sheet_data_dict)
        return balance_sheet_data_dict
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading balance sheet failed.')
        pass
        
def save_cash_flow_price_derived_to_dict(worksheet, start_cell, end_cell):
    try:
        cash_flow_data_dict = {}
        
        # Extract values from each cell in the specified range
        for row in worksheet[start_cell:end_cell]:
            key = row[0].value
            if key == None:
                print('[INFO] None type value, no key exist')
                continue
            else:
                  # Extract the key from the first cell of the row
                values = [cell.value for cell in row[1:] if cell.value != None]  # Extract values from the remaining cells of the row
                if values == None:
                    print('[INFO] None type value, no value exist')
                    continue
                if values != None:
                    cash_flow_data_dict[key] = values
                else:
                    continue
        
        print('[INFO] cahs flow data saved in dictionary')
        print(cash_flow_data_dict)
        return cash_flow_data_dict
    except (RuntimeError, TypeError, NameError):
        print(f'[INFO] Loading cash flow failed.')
        pass

def get_data_from_data_sheet(workbook):
    sheet_DataSheet = load_DataSheet(workbook)
    company_name = get_company_name(sheet_DataSheet)
    meta_data_dict = save_metadata_to_dict(sheet_DataSheet, 'A6', 'B9')
    profit_and_loss_data_dict = save_profit_loss_to_dict(sheet_DataSheet, 'A16', 'K31')
    quarter_data_dict = save_quarters_to_dict(sheet_DataSheet, 'A41', 'K50')
    balance_sheet_dict = save_balance_sheet_to_dict(sheet_DataSheet, 'A56', 'K72' )
    cash_flow_price_derived_dict = save_cash_flow_price_derived_to_dict(sheet_DataSheet, 'A81', 'K93')
    print(sheet_DataSheet.title)
    print('\n\n')
    print(company_name)
    print('\n\n')
    print(meta_data_dict)
    print('\n\n')
    print(profit_and_loss_data_dict)
    print('\n\n')
    print(quarter_data_dict)
    print('\n\n')
    print(balance_sheet_dict)
    print('\n\n')
    print(cash_flow_price_derived_dict)
    print('\n\n')
    return company_name, meta_data_dict, profit_and_loss_data_dict, quarter_data_dict, balance_sheet_dict, cash_flow_price_derived_dict
  

'''    
wb = load_workbook('./Tata_motors.xlsx', rich_text=True)
print(wb.sheetnames)

sheet_DataSheet = wb['Data Sheet']

save_metadata_to_dict(sheet_DataSheet, 'A6', 'B9')   
save_profit_loss_to_dict(sheet_DataSheet, 'A16', 'K31')
'''

